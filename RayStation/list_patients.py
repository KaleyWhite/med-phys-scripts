import clr
from datetime import datetime
import os
import sys
from typing import Dict, List, Optional, Tuple

from bidict import bidict  # Two-way dictionary

from connect import *
from connect.connect_cpython import PyScriptObject  # For type hints

from openpyxl import Workbook
from openpyxl.styles import numbers

clr.AddReference('System')
clr.AddReference('System.Drawing')
clr.AddReference('System.Windows.Forms')
from System import EventArgs  # For type hints
from System.Drawing import *
from System.Windows.Forms import *


# Absolute path of the directory to create the output file in
# Directory does not have to already exist
OUTPUT_DIR = os.path.join('T:', os.sep, 'Physics', 'Scripts', 'Output Files', 'List Patients')


def get_tx_technique(beam_set: PyScriptObject) -> str:
    """Determines the treatment technique of the beam set

    Code modified from RS support
    
    Arguments
    ---------
    beam_set: The beam set whose treatment technique to return

    Returns
    -------
    The treatment technique ("SMLC", "SRS", "SBRT", "SABR", "VMAT", "DMLC", "Conformal", "Conformal Arc", or "ApplicatorAndCutout"), or "[Unknown]" if the treatment technique could not be determined
    """
    if beam_set.Modality == 'Photons':
        if beam_set.PlanGenerationTechnique == 'Imrt':
            if beam_set.DeliveryTechnique == 'SMLC':
                return 'SMLC'
            if beam_set.DeliveryTechnique == 'DynamicArc':
                if beam_set.Prescription is not None and beam_set.FractionationPattern is not None:
                    fx = beam_set.FractionationPattern.NumberOfFractions
                    rx = beam_set.Prescription.PrimaryPrescriptionDoseReference.DoseValue / fx
                    if rx >= 600:
                        if fx in [1, 3]:
                            return 'SRS'
                        if fx == 5:
                            return 'SBRT'
                        if fx in [6] + list(range(8, 16)):
                            return 'SABR'
                return 'VMAT'
            if beam_set.DeliveryTechnique == 'DMLC':
                return 'DMLC'
        elif beam_set.PlanGenerationTechnique == 'Conformal':
            if beam_set.DeliveryTechnique == 'SMLC':
                # return 'SMLC' # Changed from 'Conformal'. Failing with forward plans.
                return 'Conformal'
                # return '3D-CRT'
            if beam_set.DeliveryTechnique == 'Arc':
                return 'Conformal Arc'
    elif beam_set.Modality == 'Electrons' and beam_set.PlanGenerationTechnique == 'Conformal' and beam_set.DeliveryTechnique == 'SMLC':
        return 'ApplicatorAndCutout'
    return '[Unknown]'


class ListPatientsForm(Form):
    """Form that allows the user to specify filters for selection of MRNs to write to a file

    Form does not actually write the MRNs
    
    Attributes
    ----------
    max_num_pts (float): Maximum number of MRNs that should be written to the file
    keywords (List[str]): List of keywords
    checked_values (Dict[str, List[str]]): Lists of checked values for all multiple-select filters

    Implementation details:
    The code admittedly spaghetti-ish and needs some serious restructuring, but here goes.
    For each filter, there is a checkbox and a groupbox. If the checkbox is checked, the filter is applied. If the checkbox is not checked, the filter is not applied and the groupbox is disabled. The checkboxes and groupboxes are associated with the bidict `_filter_cbs_gbs`. There is also a "Select all" checkbox for selecting/deselecting all filters.
    The keywords filter is just a textbox that the user enters keywords into, one per line. When this text is changed, the `keywords` attribute is updated.
    The other filters (sex, patient position, and treatment technique) are multiple-select. Each filter includes a "Select all" checkbox to check all options. The attribute `checked_values` has keys that are the text of the groupboxes, and values that are lists of the selected options. When the user checkes a checkbox, `checked_values` is updated.
    There is also a textbox for the maximum number of MRNs to write to the file. This is stored in the attribute `max_num_pts`. The "OK" button is enabled only if `max_num_pts` is a positive integer, or "[All]" (case insensitive) for no limit.
    If no filters are effectively applied, a message at the bottom of the form says so. Even if a filter checkbox is checked, the filter is EFFECTIVELY not applied if no values are provided (keywords filter), or either no or all options are checked (multiple-select filter).
    The "Select all" checkbox functionality is implemented with three-state checkboxes whose Tag hold their previous state.
    """
    # Absolute path of directory to save output file in
    # This directory does not have to already exist
    OUTPUT_DIR = os.path.join('T:', os.sep, 'Physics', 'Scripts', 'Output Files', 'List Patients')

    def __init__(self):
        """Initializes a ListPatientsForm object"""
        # Public attributes
        self.max_num_pts = float('inf')  # Maximum number of MRNs that should be written to a file
        self.keywords: List[str] = []  # List of keywords to check the patients for
        self.checked_values: Dict[str, List[str]] = {}  # Dictionary of <filter title> : [<checked options>] (e.g., 'Sex': ['Male', 'Other'])

        self._set_up_form()
        self._y = 15  # Vertical coordinate of next control

        # Label with script description/instructions
        text = 'MRNs of patients meeting all of the following filter criteria will be written to a file in "' + self.OUTPUT_DIR + '".\n' \
             + 'The filename includes a timestamp.\n' \
             + 'Check the filters that you want to apply.'
        l = self._add_lbl(txt=text)
        self._y += l.Height + 15

        # User input for `max_num_pts``
        self._add_num_pts_input()

        # "Select all filters" checkbox
        self._select_all_filters_cb = self._add_checkbox(is_select_all=True, txt='Select all filters')
        self._y += self._select_all_filters_cb.Height + 10

        # Associate the filters' groupboxes with their checkboxes
        self._filter_cbs_gbs: Dict[CheckBox, GroupBox] = bidict()

        # Keywords filter. `keywords` value comes from the textbox
        instrs = 'The following are searched for the keywords (case insensitive):\n    -  '
        instrs += '\n    -  '.join(['Case names, body sites, comments, and diagnoses', 'Exam names', 'Plan names and comments', 'Beam set names and comments', 'Beam names and descriptions', 'DSP names', 'Rx descriptions and structure names'])
        instrs += '\nEnter one keyword per line:'
        self._keywords_gb, gb_y = self._add_filter('Keywords', instrs)

        self._keyword_tb = TextBox()
        self._set_up_keyword_tb()
        self._keyword_tb.Location = Point(15, self._keywords_gb.Height - 15)
        self._keywords_gb.Controls.Add(self._keyword_tb)
        self._y += self._keywords_gb.Height + 10

        # Sex filter
        gb = self._add_checkboxes_filter('Sex', ['Male', 'Female', 'Other'])
        self._y += gb.Height + 10

        # Patient Position filter
        gb = self._add_checkboxes_filter('Patient Position', ['FFS', 'HFP', 'HFS'])
        self._y += gb.Height + 10

        # Tx technique filter
        gb = self._add_checkboxes_filter('Treatment Technique', ['Applicator and cutout', 'Conformal', 'Conformal Arc', 'DMLC', 'SABR', 'SBRT', 'SRS', 'SMLC', 'VMAT', '[Unknown]'])
        self._y += gb.Height + 15

        # Warning label
        # Visible if no filters (effectively) will be applied
        self._warning_lbl = self._add_lbl(txt='No filters are applied.\nAll MRNs will be written to the file.')
        self._y += self._warning_lbl.Height

        # "OK" button
        self._ok_btn = Button()
        self._set_up_ok_btn()

    # ---------------------------------------------------------------------------- #
    #                                Event handlers                                #
    # ---------------------------------------------------------------------------- #

    # --------------------------------- Textboxes -------------------------------- #

    def _num_pts_tb_TextChanged(self, sender: TextBox, event: EventArgs) -> None:
        """Handles the event of changing the contents of the textbox for the max number of patients

        If the text value is a valid positive integer, sets `max_num_pts` to that integer.
        Otherwise, if the text value is '[all]', case insensitive, sets `max_num_pts` to infinity
        Otherwise, sets `max_num_pts` to None
        
        If the new `max_num_pts` value is None and the text is not empty, set red textbox background color
        Otherwise, set white background color
        
        Enable "OK" button if `max_num_pts` is valid. Disable otherwise.
        """
        num_pts = self._num_pts_tb.Text
        # If textbox is empty, invalid value, but don't turn textbox red
        if num_pts == '':
            self._ok_btn.Enabled = False
            self._num_pts_tb.BackColor = Color.White
            self.max_num_pts = None
        # '[All]' means no limit on number of MRNs
        elif num_pts.lower() == '[all]':
            self._ok_btn.Enabled = True
            self._num_pts_tb.BackColor = Color.White
            self.max_num_pts = float('inf')
        # Neither empty string nor '[All]'
        else:
            try:
                # Valid positive integer?
                num_pts = int(num_pts)
                if num_pts > 0:
                    self._ok_btn.Enabled = True
                    self._num_pts_tb.BackColor = Color.White
                    self.max_num_pts = num_pts
                else:  # Non-numeric or -integer
                    raise ValueError
            except ValueError:
                # Invalid input
                self._ok_btn.Enabled = False
                self._num_pts_tb.BackColor = Color.Red
                self.max_num_pts = None

    def _keyword_tb_TextChanged(self, sender: TextBox, event: EventArgs) -> None:
        """Handles the event of changing the value in the textbox in the keywords filter

        If the keywords textbox is empty, empties the `keywords` list
        Otherwise, the `keywords` list elements are the lines in the textbox
        
        Shows or hides the "no filters will be applied" warning
        """
        keywords = self._keyword_tb.Text.strip()  # Remove leading and trailing whitespace
        if keywords == '':  # No keywords provided
            self.keywords = []
        else:  # Update `keywords` to the lines in the textbox
            self.keywords = [keyword.lower().strip() for keyword in keywords.split('\n')]
        self._set_warning_label()  # Show or hide the "no filters will be applied" warning

    # -------------------------------- Checkboxes -------------------------------- #

    def _select_all_checkbox_Click(self, sender: CheckBox, event: EventArgs) -> None:
        """Handles the event of clicking a "Select all" checkbox

        If the checkbox was previously unchecked or indeterminate, check it and all of the item checkboxes it is associated with
        Otherwise, uncheck it and all its item checkboxes

        If the clicked checkbox is the "Select all filters" checkbox, also enable/disable the filter groupboxes when the filter checkboxes are checked/unchecked
        Otherwise, update the filter's entry in `checked_values`

        Shows or hides the "no filters will be applied" warning
        """
        prev_state = sender.Tag == CheckState.Checked  # Checkbox tag tracks previous checkstate
        new_state = not prev_state  # We will toggle the state
        
        # Set "Select all" checkstate
        if prev_state:  # Used to be checked, so uncheck
            sender.CheckState = sender.Tag = CheckState.Unchecked
        else:  # Used to be unchecked or indeterminate, so check
            sender.CheckState = sender.Tag = CheckState.Checked

        # Set options checkstates
        parent = sender.Parent
        if isinstance(parent, Form):  # It's the "select all filters" checkbox
            # Check/uncheck the filter checkboxes and enable/disable the filter groupboxes
            for cb, gb in self._filter_cbs_gbs.items():
                cb.Checked = gb.Enabled = new_state  # Enable groupbox only if checkbox is checked
        else:  # It's a checkbox within a filter
            item_cbs = [ctrl for ctrl in parent.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is None]  # All checkboxes in the groupbox that are not "Select all"
            for cb in item_cbs:
                cb.Checked = new_state  # Toggle checkstate
            # Update `checked_values`
            gb_key = parent.Text[:-1]  # Remove colon from groupbox text
            if new_state:  # Populate the values list with all checkbox texts
                self.checked_values[gb_key] = [cb.Text for cb in item_cbs]
            else:  # Empty the values list if the checkboxes were unchecked
                self.checked_values[gb_key] = []
        self._set_warning_label()  # Show/hide the "no filters will be applied" warning

    def _item_checkbox_Click(self, sender: CheckBox, event: EventArgs) -> None:
        """Handles the event that is clicking a checkbox that is not "Select all"

        If all non-"Select all" checkboxes in the group are now checked, check the "Select all" checkbox
        Otherwise, if none of the checkboxes are now checked, uncheck the "Select all" checkbox
        Otherwise, set the "Select all" checkstate to indeterminate

        Updates `checked_values` and shows/hides the "no filters will be applied" warning
        """
        parent = sender.Parent
        select_all = next(ctrl for ctrl in parent.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is not None)  # Get the "Select all" checkbox in this group
        if isinstance(parent, Form):  # A filter checkbox was clicked
            gb = self._filter_cbs_gbs[sender]  # Corresponding groupbox
            gb.Enabled = sender.Checked  # Enable groupbox only if checkbox is now checked
        else:  # The checkbox is inside a filter
            gb = parent
        
        item_cbs = [ctrl for ctrl in parent.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is None]  # All checkboxes in the groupbox that are not "Select all"
        item_cbs_checked = [cb for cb in item_cbs if cb.Checked]  # All checked checkboxes that are not "Select all"
        
        # Check or uncheck "Select all"
        if len(item_cbs) == len(item_cbs_checked):  # All checkboxes are checked
            select_all.CheckState = select_all.Tag = CheckState.Checked  # Check "Select all"
        elif item_cbs_checked:  # Some checkboxes are checked, some unchecked
            select_all.CheckState = select_all.Tag = CheckState.Indeterminate  # "Select all" checkstate is indeterminate
        else:  # All checkboxes are unchecked
            select_all.CheckState = select_all.Tag = CheckState.Unchecked  # Uncheck "Select all"
        
        # Update `checked_values` as the text of the checked checkboxes
        if gb is parent:
            self.checked_values[parent.Text[:-1]] = [cb.Text for cb in item_cbs_checked]
        
        self._set_warning_label()  # Warning label visibility depends on whether any filters are applied

    def _ok_btn_Click(self, sender, event):
        # Event handler for clicking the 'OK' button
        
        self.DialogResult = DialogResult.OK

    # ---------------------------------------------------------------------------- #
    #                                 Styling/setup                                #
    # ---------------------------------------------------------------------------- #

    def _set_up_form(self) -> None:
        """Styles this Form"""
        self.Text = 'List Patients'  # Form title
        
        # Adapt form size to controls
        self.AutoSize = True  
        self.AutoSizeMode = AutoSizeMode.GrowAndShrink
        self.MinimumSize = Size(TextRenderer.MeasureText(self.Text, SystemFonts.CaptionFont).Width + 100, 0)  # At least as wide as title plus room for "X" button, etc.

        self.FormBorderStyle = FormBorderStyle.FixedToolWindow  # User cannot minimize, maximize, or resize form, but they can cancel ('X out of') it
        self.StartPosition = FormStartPosition.CenterScreen  # Position form in middle of screen

    def _set_up_keyword_tb(self) -> None:
        """Styles the keywords textbox"""
        self._keyword_tb.AutoSize = True
        self._keyword_tb.MinimumSize = Size(150, 75)
        self._keyword_tb.Multiline = True  # Keywords are separated by newlines
        self._keyword_tb.TextChanged += self._keyword_tb_TextChanged

    def _set_up_ok_btn(self) -> None:
        """Styles the "OK" button and adds it to the Form"""
        self._ok_btn.Location = Point(self.ClientSize.Width - 50, self._y)  # Right align
        self._ok_btn.Text = 'OK'
        self._ok_btn.Click += self._ok_btn_Click
        self.Controls.Add(self._ok_btn)

    # ---------------------------------------------------------------------------- #
    #                                 Add Controls                                 #
    # ---------------------------------------------------------------------------- #

    # ------------------------------ Basic controls ------------------------------ #

    def _add_checkbox(self, **kwargs) -> Checkbox:
        """Adds a checkbox to a control

        Keyword Arguments
        -----------------
        parent: Control to add the checkbox to
                Defaults to self
        x: x-coordinate of the checkbox
           Defaults to 15
        y: y-coordinate of the checkbox
           Defaults to the parent's `_y` attribute, if applicable, otherwise 15
        is_select_all: True if the checkbox is "Select all", False otherwise
                       Defaults to False
        txt: Text of the checkbox
             Defaults to "Select all" for a "Select all" checkbox, the empty string otherwise

        Returns
        -------
        The new checkbox
        """
        # Extract keyword args
        parent = kwargs.get('parent', self)
        x = kwargs.get('x', 15)
        y = kwargs.get('y', parent._y if hasattr(parent, '_y') else 15)
        is_select_all = kwargs.get('is_select_all', False)
        txt = kwargs.get('txt', 'Select all' if is_select_all else '')

        # Create checkbox
        cb = CheckBox()
        cb.AutoSize = True
        cb.AutoSizeMode = AutoSizeMode.GrowAndShrink
        if is_select_all:
            cb.IsThreeState = True
            cb.CheckState = cb.Tag = CheckState.Unchecked
            cb.Click += self._select_all_checkbox_Click
        else:
            cb.Checked = False
            cb.Click += self._item_checkbox_Click
        cb.Location = Point(x, y)
        cb.Text = txt
        parent.Controls.Add(cb)
        return cb

    def _add_groupbox(self, **kwargs) -> None:
        """Adds a groupbox to the form

        Keyword Arguments
        -----------------
        x: x-coordinate of the groupbox
           Defaults to 15
        txt: Text of the checkbox
             Defaults to the empty string
             If provided, a colon is appended

        Returns
        -------
        The new groupbox
        """
        # Extract keyword arguments
        x = kwargs.get('x', 15)
        txt = kwargs.get('txt', '')
        if txt != '':
            txt += ':'

        # Create groupbox
        gb = GroupBox()
        gb.Enabled = False
        gb.AutoSize = True
        gb.AutoSizeMode = AutoSizeMode.GrowAndShrink
        gb.MinimumSize = Size(TextRenderer.MeasureText(txt, gb.Font).Width + 50, 0)
        gb.Location = Point(50, self._y)
        gb.Text = txt
        self.Controls.Add(gb)
        return gb

    def _add_lbl(self, **kwargs) -> Label:
        """Adds a label to a control

        Arguments
        ---------
        lbl_txt: Text of the Label

        Keyword Arguments
        -----------------
        parent: Control to add the label to
        x: x-coordinate of new label
           Defaults to 15
        y: y-coordinate of the checkbox
           Defaults to the parent's `_y` attribute, if applicable, otherwise 15
        bold: True if label text should be bold, False otherwise
        txt: Text of the label
             Defaults to the empty string

        Returns
        -------
        The label
        """
        # Extract keyword arguments
        parent = kwargs.get('parent', self)
        x = kwargs.get('x', 15)
        y = kwargs.get('y', parent._y if hasattr(parent, '_y') else 15)
        bold = kwargs.get('bold', False)
        txt = kwargs.get('txt', '')

        # Create label
        l = Label()
        l.AutoSize = True
        l.AutoSizeMode = AutoSizeMode.GrowAndShrink
        if bold:
            l.Font = Font(l.Font, FontStyle.Bold)
        l.Location = Point(x, y)
        l.Text = txt
        parent.Controls.Add(l)
        return l

    # -------------------------------- User inputs ------------------------------- #

    def _add_num_pts_input(self) -> None:
        """Adds controls to allow the user to input the maximum number of MRNs that should be written to a file"""
        l = self._add_lbl(txt='Max # of MRN(s) to write:')

        self._num_pts_tb = TextBox()
        self._num_pts_tb.Location = Point(20 + l.Width, self._y)  # Immediately to the right of the label
        self._num_pts_tb.Text = '[All]'
        self._num_pts_tb.Width = 25
        self._num_pts_tb.TextChanged += self._num_pts_tb_TextChanged
        self.Controls.Add(self._num_pts_tb)
        self._y += 50

    def _add_filter(self, name: str, instrs: Optional[str] = None) -> Tuple[GroupBox, int]:
        """Adds a checkbox and groupbox for a filter

        Arguments
        ---------
        name: Title of the filter. Used as text of the groupbox (plus a colon)
        instrs: Instructions to display to the user
                If None, no instructions are displayed
        
        Returns
        -------
        A 2-tuple: the groupbox and the y-coordinate for the next control in the groupbox
        """
        # Checkbox
        cb = self._add_checkbox(x=30)

        ## Groupbox
        gb = self._add_groupbox(x=50, txt=name)
        gb_y = 15  # Vertical coordinate of next control in groupbox

        # Filter description/instructions
        if instrs is not None:
            l = self._add_lbl(parent=gb, y=gb_y, txt=instrs)
            gb_y += l.Height

        self._filter_cbs_gbs[cb] = gb  # Associate the checkbox and groupbox

        return gb, gb_y

    def _add_checkboxes_filter(self, name: str, items: List[str], instrs: Optional[str] = None) -> GroupBox:
        """Adds a filter that is multiple-select (checkboxes)

        Arguments
        ---------
        name: Title of the filter. Used as text of the groupbox (plus a colon)
        items: List of checkbox options
        instrs: Instructions to display to the user
                If None, no instructions are displayed

        Returns
        -------
        The groupbox part of the filter
        """
        gb, gb_y = self._add_filter(name, instrs)  # Add checkbox and groupbox
        
        # "Select all" checkbox
        cb = self._add_checkbox(parent=gb, y=gb_y, is_select_all=True)
        gb_y += cb.Height

        # Item checkboxes
        for item in items:
            cb = self._add_checkbox(parent=gb, x=30, y=gb_y, txt=item)
            gb_y += cb.Height

        self.checked_values[name] = []  # By default, no items are checked

        return gb

    # ---------------------------------------------------------------------------- #
    #                                 Warning label                                #
    # ---------------------------------------------------------------------------- #

    def _set_warning_label(self) -> None:
        """Shows or hides a warning label that states that no filters will be applied
        
        Effectively, a filter is not applied if its checkbox is unchecked (obviously), or if no keywords are provided (for keywords filter) or no or all checkboxes are checked (for other filters)
        """
        keywords_applied = self._filter_cbs_gbs.inverse[self._keywords_gb].Checked and self.keywords  # Keywords filter checkbox is checked and `keyords` list is nonempty
        if keywords_applied:
            self._warning_lbl.Visible = False
        else:  # Are any multi-select (checkbox) filters applied?
            for cb, gb in self._filter_cbs_gbs.items():
                if not cb.Checked or gb is self._keywords_gb:  # Unchecked filters are not applied, so ignore. Skip keywords filter, which has already been checked
                    continue
                vals = self.checked_values[gb.Text[:-1]]  # Strip colon of groupbox text to get key of `checked_values`
                num_cbs = sum(1 for ctrl in gb.Controls if isinstance(ctrl, CheckBox) and ctrl.Tag is None)  # Number of checkboxes in the groupbox that are not "Select all"
                if 0 < len(vals) < num_cbs:  # Some, but not all, checkboxes are checked
                    self._warning_lbl.Visible = False
                    return
            self._warning_lbl.Visible = True  # No checkbox filters are effectively applied


def list_patients():
    """Writes an Excel file with the MRNs of RayStation patients matching all the user-selected filters

    Filters:
    - Keyword(s): Any of the provided keywords is part of any of the following:
        * A case name, body site, comment, or dx
        * A plan name or comment
        * An exam name
        * A beam set name or comment
        * A beam name or description
        * A DSP name
        * An Rx description or structure name
    - Sex(es)
    - Patient position(s): The patient position of an exam is as specified. If keywords are also provided, keyword matches associated with a plan must be for a planning exam with tone of these patient positions.
    - Treatment technique(s): The treatment delivery technique of a beam set. If keywords are also provided, kewyprd matches associated with a beam set must be for a beam set with one of these tx techniques.

    Filename is the datetime and the number of matching patients
    Note that patients that are currently open will not be checked!
    If no patients are found, file contains a single line: 'No matching patients.'
    """

    # Get filters from user
    form = ListPatientsForm()
    form.ShowDialog()
    if form.DialogResult != DialogResult.OK:  # "OK" button was never clicked (user probably exited GUI)
        sys.exit()

    max_num_pts = form.max_num_pts
    keywords = form.keywords
    sex = form.checked_values['Sex']
    pt_pos = form.checked_values['Patient Position']
    tx_techniques = form.checked_values['Treatment Technique']

    patient_db = get_current('PatientDB')

    matching_mrns = []  # MRNs to write to output file
    all_pts = patient_db.QueryPatientInfo(Filter={})  # Get all patient info in RS DB
    for pt in all_pts:  # Iterate over all patients in RS
        if len(matching_mrns) == max_num_pts:  # We have reached the user-defined limit # of pts
            break

        try:
            pt = patient_db.LoadPatient(PatientInfo=pt)  # Load the patient
        except:  # Someone has the patient open
            continue
 
        # If sex filter applied and patient doesn't match, move on to next patient
        if sex and pt.Gender not in sex:
            continue
        
        if keywords:  # Keywords filter is applied
            names_to_chk = []  # Strings to check whether they contain any of the keywords
            for case in pt.Cases:  # Check all cases
                for exam in case.Examinations:  # Check all exams
                    if not pt_pos or exam.PatientPosition in pt_pos:  # Check this exam if either (a) patient position filter is not applied, or (b) exam patient position matches the filter
                        names_to_chk.extend([case.BodySite, case.CaseName, case.Comments, case.Diagnosis])  # We will check certain case attributes for the keywords
                        names_to_chk.append(exam.Name)  # We will check the exam name for the keywords
                        for plan in case.TreatmentPlans:  # Check each plan
                            if plan.GetTotalDoseStructureSet().OnExamination.Name.lower() == exam.Name.lower():  # The plan must be on the current exam
                                names_to_chk.extend([plan.Name, plan.Comments])  # We will check certain plan attributes for the keywords
                                for bs in plan.BeamSets:  # Check each beam set
                                    if not tx_techniques or get_tx_technique(bs) in tx_techniques:  # Check this beam set if either (a) the treatment technique filter is not applied, or (b) the beam set's treatment technique matches the filter
                                        names_to_chk.extend([bs.Comment, bs.DicomPlanLabel])  # We will check certain beam set attributes for the keywords
                                        for b in bs.Beams:  # Check each beam
                                            names_to_chk.extend([b.Description, b.Name])  # We will check certain beam attributes for the keywords
                                        for dsp in bs.DoseSpecificationPoints:  # Check each DSP
                                            names_to_chk.append(dsp.Name)  # We will check DSP names for the keywords
                                        if bs.Prescription is not None:
                                            # We will check certain Rx attributes for the keywords
                                            names_to_chk.append(bs.Prescription.Description)
                                            for rx in bs.Prescription.PrescriptionDoseReferences:
                                                if hasattr(rx, 'OnStructure'):
                                                    names_to_chk.append(rx.OnStructure.Name)
           
            # Check each name for each keyword (case insensitive)
            names_to_chk = set(name.lower() for name in names_to_chk if name is not None)  # Remove duplicates and None from names list. Standardize keywords to lowercase.
            keyword_match = False  # Assume no match
            for keyword in keywords:
                for name in names:
                    if keyword in name and (not keyword.startswith('pelvi') or 'abd' not in name):  # Very special case: if searching for "pelvis", we don't want, e.g, "abdomen/pelvis"
                        keyword_match = True
                        break
            if not keyword_match:  # If no keywords present in any of the names, move on to next patient
                continue
        
        # Keywords filter is not applied, so exam with correct patient position doesn't have to be associated with a matching keyword
        elif pt_pos and not any(exam.PatientPosition in pt_pos for case in pt.Cases for exam in case.Examinations):
            continue

        elif tx_techniques and not any(get_tx_technique(bs) in tx_techniques for case in pt.Cases for plan in case.TreatmentPlans for bs in plan.BeamSets):
            continue
        
        matching_mrns.append(pt.PatientID)

    ## Write output file
    wb = Workbook()  # Create Excel data
    ws = wb.active
    
    # Format column as text, not number
    for row in ws:
        row[1].number_format = numbers.FORMAT_TEXT
    
    # Write data to file
    if matching_mrns:  # There are patients matching all criteria -> write in sorted order, one MRN per line
        for i, mrn in enumerate(sorted(matching_mrns)):
            ws.cell(i + 1, 1).value = mrn
    else:  # No matching patients, so say so in the first cell of the output file.
        ws.cell(1, 1).value = 'No matching patients.'

    ## Write Excel data to file
    # Create output directory if it doesn't already exist
    if not os.path.isdir(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    # Construct filename
    dt = datetime.now().strftime('%Y-%m-%d %H_%M_%S')
    filename = dt + ' (' + str(len(matching_mrns)) + ' Patients).xlsx'
    # Write to file
    wb.save(os.path.join(OUTPUT_DIR, filename))
